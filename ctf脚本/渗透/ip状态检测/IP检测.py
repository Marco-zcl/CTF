import asyncio
import aiohttp
from openpyxl import Workbook
from bs4 import BeautifulSoup
import re
import uuid
from tqdm import tqdm

async def get_title(session, url, index, ws):
    try:
        async with session.get(url) as response:
            if response.status == 200:
                html = await response.text()
                soup = BeautifulSoup(html, 'html.parser')
                title_tag = soup.title
                if title_tag:
                    title = title_tag.string.strip()
                else:
                    title = "Null"
            else:
                title = str(response.status)
    except Exception as e:
        title = "500"

    ws.cell(row=index, column=3).value = title

async def process_urls(input_file_name, output_file_name):
    async with aiohttp.ClientSession() as session:
        with open(input_file_name, 'r') as input_file:
            urls = input_file.readlines()

        wb = Workbook()
        ws = wb.active
        ws.append(["URL", "Status", "Title"])

        tasks = []
        for index, url in enumerate(urls, start=2):  # 开始的行号从2开始，因为第一行是标题
            url = url.strip()
            ws.cell(row=index, column=1).value = url
            task = asyncio.create_task(get_title(session, url, index, ws))
            tasks.append(task)

        print("开始运行程序，正在RUN请不要关闭")
        # 显示进度条
        for _ in tqdm(asyncio.as_completed(tasks), total=len(tasks), desc="运行中", unit="tasks"):
            await _

        output_file_name = str(uuid.uuid4()) + ".xlsx"  # 随机生成文件名
        wb.save(output_file_name)

        return output_file_name

async def main():
    input_file_name = input("请输入要读取的文件名: ")
    output_file_name = "output.xlsx"  # 初始化 output_file_name 变量

    output_file_name = await process_urls(input_file_name, output_file_name)

    print(f"处理完成，结果已保存到 {output_file_name} 文件中。")

if __name__ == "__main__":
    asyncio.run(main())
